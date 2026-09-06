from __future__ import annotations

import shutil
import sys
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

from app.config import AppConfig, save_config, load_config  # noqa: E402
from app.dates import parse_date  # noqa: E402
from app.domain import is_closed, is_open, is_overdue, is_placed, is_status_open  # noqa: E402
from app.excel.service import ExcelService  # noqa: E402
from app.stats import kpis  # noqa: E402
from app.validation import validate_work_order  # noqa: E402


@pytest.fixture()
def workbook(tmp_path):
    src = ROOT / "file.xlsx"
    dest = tmp_path / "file.xlsx"
    shutil.copy2(src, dest)
    cfg = AppConfig()
    cfg.excel_path = str(dest)
    cfg.backup_dir = str(tmp_path / "backups")
    save_config(cfg)
    svc = ExcelService()
    svc.invalidate()
    yield dest, svc
    cfg = AppConfig()
    cfg.excel_path = str(ROOT / "file.xlsx")
    cfg.backup_dir = str(ROOT / "backups")
    save_config(cfg)


def test_excel_exists():
    assert (ROOT / "file.xlsx").exists()


def test_read_work_orders(workbook):
    _, svc = workbook
    recs = svc.get_all(force=True)
    assert len(recs) >= 2000
    rec = recs[0]
    assert rec["work_order_id"]
    assert rec["record_id"]
    assert rec["status"]
    assert rec["_sheet"] in {"Linkco_MR_Log (SH5 & SH1)", "Linkco_MR_Log (F5)"}
    assert rec["department"] in {"SH5-SH1", "F5"}
    assert parse_date(rec["created_date"]) is not None or rec["created_date"] == ""


def test_unique_record_ids(workbook):
    _, svc = workbook
    recs = svc.get_all(force=True)
    ids = [r["record_id"] for r in recs]
    assert len(ids) == len(set(ids))
    # IM WO # may repeat (multiple MRs per work order)
    assert len({r["work_order_id"] for r in recs}) < len(recs)


def test_kpis_match_records(workbook):
    _, svc = workbook
    recs = svc.get_all(force=True)
    k = kpis(recs)
    assert k["total"] == len(recs)
    assert k["closed"] == sum(1 for r in recs if is_closed(r))
    assert k["open"] == sum(1 for r in recs if is_status_open(r))
    assert k["placed"] == sum(1 for r in recs if is_placed(r))
    assert k["open"] + k["closed"] <= k["total"]
    assert k["open"] + k["placed"] + k["closed"] <= k["total"]
    assert k["overdue"] == sum(1 for r in recs if is_overdue(r))
    assert 0 <= k["completion_rate"] <= 100


def test_update_writes_excel(workbook):
    path, svc = workbook
    recs = svc.get_all(force=True)
    target = next(r for r in recs if is_open(r))
    rid = target["record_id"]
    updated = svc.update_record(rid, {"remarks": "Updated by automated test"}, username="pytest")
    assert updated["remarks"] == "Updated by automated test"
    again = svc.get_by_id(rid)
    assert again["remarks"] == "Updated by automated test"


def test_create_appends_row(workbook):
    _, svc = workbook
    before = len(svc.get_all(force=True))
    created = svc.create_record(
        {
            "description": "Test create from pytest",
            "status": "OPEN",
            "priority": "MEDIUM",
            "department": "F5",
            "work_type": "Local PO",
        },
        username="pytest",
    )
    assert created["work_order_id"]
    recs = svc.get_all(force=True)
    assert len(recs) == before + 1
    assert sum(1 for r in recs if r["record_id"] == created["record_id"]) == 1
    created_d = parse_date(created["created_date"])
    due_d = parse_date(created["due_date"])
    assert created_d and due_d
    assert (due_d.date() - created_d.date()).days == 5


def test_validation_close_before_create():
    errors = validate_work_order(
        {
            "description": "x",
            "status": "CLOSED",
            "created_date": "2026-09-05 10:00",
            "closed_date": "2026-09-01 10:00",
        }
    )
    assert any("Closing date" in e for e in errors)


def test_auth_and_dashboard(workbook):
    from app.main import app
    from app import database

    database.init_db()
    client = TestClient(app)
    bad = client.post("/api/auth/login", json={"username": "admin", "password": "wrong"})
    assert bad.status_code == 401
    ok = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})
    assert ok.status_code == 200
    token = ok.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}
    dash = client.get("/api/dashboard", headers=headers)
    assert dash.status_code == 200
    body = dash.json()
    assert body["kpis"]["total"] > 2000
    assert body["mindmap"]["root"]["value"] == body["kpis"]["total"]
    assert body["mindmap"]["branches"]
    assert body.get("recent") is not None
    assert body.get("options")
    assert body.get("groups", {}).get("status")
    assert body.get("groups", {}).get("supplier")
    again = client.get("/api/dashboard", headers=headers)
    assert again.status_code == 200
    layout = client.put("/api/auth/layout", headers=headers, json={"widgets": [{"id": "kpis_today", "type": "kpis_today", "span": "full"}]})
    assert layout.status_code == 200
    got = client.get("/api/auth/layout", headers=headers)
    assert got.status_code == 200
    assert got.json()["widgets"][0]["type"] == "kpis_today"
    wos = client.get("/api/work-orders?page_size=10", headers=headers)
    assert wos.status_code == 200
    assert wos.json()["total"] > 2000
    # Empty-filter dashboard used to request /api/work-orders&page_size=… (404).
    glued = client.get("/api/work-orders&page_size=8&sort=created_date", headers=headers)
    assert glued.status_code == 404
    ok_list = client.get("/api/work-orders?page_size=8&sort=created_date", headers=headers)
    assert ok_list.status_code == 200
    assert len(ok_list.json()["items"]) == 8
    health = client.get("/api/health")
    assert health.status_code == 200


def test_upload_scans_workbook(workbook):
    from app.main import app
    from app import database
    from app.excel.service import excel_service

    database.init_db()
    excel_service.invalidate()
    path, _ = workbook
    client = TestClient(app)
    token = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"}).json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}
    content = path.read_bytes()
    res = client.post(
        "/api/sync/upload",
        headers=headers,
        files={"file": ("file.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["ok"] is True
    assert body["sync"]["record_count"] > 2000
    assert body["mindmap"]["root"]["value"] == body["sync"]["record_count"]
    reject = client.post(
        "/api/sync/upload",
        headers=headers,
        files={"file": ("notes.txt", b"not-an-excel-file", "text/plain")},
    )
    assert reject.status_code == 400


def test_ping_and_week_filter(workbook):
    from app.main import app
    from app import database
    from app.excel.service import excel_service

    database.init_db()
    excel_service.invalidate()
    client = TestClient(app)
    token = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"}).json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}
    ping = client.get("/api/sync/ping", headers=headers)
    assert ping.status_code == 200
    body = ping.json()
    assert body["synchronized"] is True
    assert body["record_count"] > 2000
    assert body["sync_token"]
    again = client.get("/api/sync/ping", headers=headers)
    assert again.json()["sync_token"] == body["sync_token"]
    week = client.get("/api/work-orders?year=2026&week=36&page_size=5", headers=headers)
    assert week.status_code == 200
    weekly = client.get("/api/dashboard/weekly?year=2026&week=36", headers=headers)
    assert weekly.status_code == 200
    assert weekly.json()["kpis"]["total"] == week.json()["total"]


def test_ops_queue_and_suppliers(workbook):
    from app.main import app
    from app import database
    from app.excel.service import excel_service
    from app.domain import is_ntp, is_open, is_overdue, is_pending_po

    database.init_db()
    excel_service.invalidate()
    _, svc = workbook
    recs = svc.get_all(force=True)
    client = TestClient(app)
    token = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"}).json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}
    queue = client.get("/api/ops/queue", headers=headers)
    assert queue.status_code == 200
    body = queue.json()
    assert body["counts"]["overdue"] == sum(1 for r in recs if is_overdue(r))
    assert body["counts"]["ntp"] == sum(1 for r in recs if is_ntp(r) and is_open(r))
    assert "overdue" in body["queues"]
    ntp = client.get("/api/work-orders?flag=ntp&page_size=5", headers=headers)
    assert ntp.status_code == 200
    assert ntp.json()["total"] == body["counts"]["ntp"]
    suppliers = client.get("/api/ops/suppliers", headers=headers)
    assert suppliers.status_code == 200
    sbody = suppliers.json()
    assert sbody["kpis"]["pending_po"] == sum(1 for r in recs if is_pending_po(r))
    assert sbody["suppliers"]
    pending = client.get("/api/work-orders?flag=pending_po&page_size=5", headers=headers)
    assert pending.status_code == 200
    assert pending.json()["total"] == sbody["kpis"]["pending_po"]


def test_preserve_other_sheets(workbook):
    from openpyxl import load_workbook
    from openpyxl.worksheet.formula import ArrayFormula

    path, svc = workbook
    recs = svc.get_all(force=True)
    target = recs[0]
    svc.update_record(target["record_id"], {"remarks": "sheet-preserve"}, username="pytest")
    wb = load_workbook(path)
    assert "Linkco_MR_Log (SH5 & SH1)" in wb.sheetnames
    assert "Linkco_MR_Log (F5)" in wb.sheetnames
    assert "SH1 & SH5 - REPORT" in wb.sheetnames
    ws = wb["Linkco_MR_Log (SH5 & SH1)"]
    assert str(ws["L1"].value).startswith("=")
    assert str(ws["A4"].value).startswith("=")
    assert isinstance(ws["H4"].value, ArrayFormula) or str(ws["H4"].value).startswith("=")
    wb.close()


def test_due_offsets_by_purchase_type(workbook):
    from datetime import timedelta

    _, svc = workbook
    recs = svc.get_all(force=True)
    expected = {
        "direct cash": 3,
        "local po": 5,
        "international": 10,
        "service": 10,
        "consumable": 2,
        "emergency": 0,
        "under warranty": 10,
        "alternative": 10,
    }
    checked = 0
    for rec in recs:
        ptype = str(rec.get("work_type") or "").strip().lower()
        if ptype not in expected:
            continue
        created = parse_date(rec.get("created_date"))
        due = parse_date(rec.get("due_date"))
        if not created or not due:
            continue
        assert due.date() == (created + timedelta(days=expected[ptype])).date()
        checked += 1
        if checked >= 40:
            break
    assert checked >= 10


def test_roles_guest_and_catalog(workbook):
    import uuid

    from app.main import app
    from app import database
    from app.excel.service import excel_service
    from app.stats import invalidate_dash_cache

    database.init_db()
    excel_service.invalidate()
    invalidate_dash_cache()
    client = TestClient(app)
    admin = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"}).json()
    headers = {"Authorization": f"Bearer {admin['access_token']}"}
    me = admin["user"]
    assert "view" in me["permissions"]
    guest_name = f"guestkpi_{uuid.uuid4().hex[:8]}"
    reader_name = f"readerkpi_{uuid.uuid4().hex[:8]}"
    created = client.post(
        "/api/users",
        headers=headers,
        json={
            "username": guest_name,
            "password": "guest123",
            "role": "guest",
            "extra_permissions": ["queue", "open"],
        },
    )
    assert created.status_code == 200, created.text
    guest = client.post("/api/auth/login", json={"username": guest_name, "password": "guest123"})
    assert guest.status_code == 200
    guser = guest.json()["user"]
    assert guser["role"] == "guest"
    assert "queue" in guser["permissions"]
    assert "open" in guser["permissions"]
    assert "users" not in guser["permissions"]
    gheaders = {"Authorization": f"Bearer {guest.json()['access_token']}"}
    forbidden = client.get("/api/users", headers=gheaders)
    assert forbidden.status_code == 403
    viewed = client.get("/api/work-orders?flag=open&page_size=5", headers=gheaders)
    assert viewed.status_code == 200
    readonly = client.post(
        "/api/users",
        headers=headers,
        json={"username": reader_name, "password": "readonly123", "role": "readonly"},
    )
    assert readonly.status_code == 200
    assert "edit" not in readonly.json()["item"]["permissions"]
    supplier = client.post("/api/catalog/suppliers", headers=headers, json={"name": "Test Supplier KPI"})
    assert supplier.status_code == 200, supplier.text
    listed = client.get("/api/catalog/suppliers", headers=headers)
    assert listed.status_code == 200
    assert any(n.lower() == "test supplier kpi" for n in listed.json()["names"])
    recs_open = client.get("/api/work-orders?flag=open&page_size=1", headers=headers)
    recs_placed = client.get("/api/work-orders?flag=placed&page_size=1", headers=headers)
    assert recs_open.status_code == 200
    assert recs_placed.status_code == 200
    dash = client.get("/api/dashboard", headers=headers)
    k = dash.json()["kpis"]
    assert recs_open.json()["total"] == k["open"]
    assert recs_placed.json()["total"] == k["placed"]
    refresh = client.post("/api/sync/refresh", headers=headers)
    assert refresh.status_code == 200
    assert refresh.json().get("hard") is True


def test_delay_fields_roundtrip_excel(workbook):
    from openpyxl import load_workbook
    from app import database

    path, svc = workbook
    recs = svc.get_all(force=True)
    target = next(r for r in recs if r.get("status"))
    rid = target["record_id"]
    updated = svc.update_record(
        rid,
        {
            "delay_kind": "placement",
            "delay_source": "site",
            "delay_justification": "Waiting on drawings",
        },
        username="pytest",
    )
    assert updated["delay_kind"] == "placement"
    assert updated["delay_source"] == "site"
    assert updated["delay_justification"] == "Waiting on drawings"
    again = svc.get_by_id(rid)
    assert again["delay_kind"] == "placement"

    wb = load_workbook(path)
    ws = wb[again["_sheet"]]
    assert ws["A3"].value == "SN"
    assert ws["T3"].value == "Link Path"
    assert ws["U3"].value == "Delay Type"
    assert ws["V3"].value == "Delay Source"
    assert ws["W3"].value == "Delay Justification"
    row = int(again["_row"])
    assert ws.cell(row, 21).value == "placement"
    assert ws.cell(row, 22).value == "site"
    assert ws.cell(row, 23).value == "Waiting on drawings"
    wb.close()

    database.upsert_record_extra(
        rid,
        "pytest",
        work_order_id=str(again["work_order_id"]),
        delay_kind="delivery",
        delay_source="supplier",
        delay_justification="stale sqlite",
    )
    result = svc.reconcile_overlay("pytest")
    extra = database.get_record_extra(rid)
    assert extra["delay_kind"] == "placement"
    assert extra["delay_source"] == "site"
    assert extra["delay_justification"] == "Waiting on drawings"
    assert result["record_count"] >= 2000


def test_reconcile_migrates_sqlite_delay_into_new_columns(workbook):
    from openpyxl import load_workbook
    from app import database
    from app.main import app
    from app.excel.service import excel_service
    from fastapi.testclient import TestClient

    path, svc = workbook
    recs = svc.get_all(force=True)
    target = recs[0]
    rid = target["record_id"]
    database.upsert_record_extra(
        rid,
        "pytest",
        work_order_id=str(target["work_order_id"]),
        delay_kind="delivery",
        delay_source="procurement",
        delay_justification="Customs hold",
    )
    result = svc.reconcile_overlay("pytest")
    assert result["wrote_excel"] is True
    assert result["pushed_to_excel"] >= 1
    again = svc.get_by_id(rid)
    assert again["delay_kind"] == "delivery"
    assert again["delay_source"] == "procurement"
    assert again["delay_justification"] == "Customs hold"
    wb = load_workbook(path)
    for sheet_name in ("Linkco_MR_Log (SH5 & SH1)", "Linkco_MR_Log (F5)"):
        ws = wb[sheet_name]
        assert ws["A3"].value == "SN"
        assert ws["T3"].value == "Link Path"
        assert ws["U3"].value == "Delay Type"
        assert ws["V3"].value == "Delay Source"
        assert ws["W3"].value == "Delay Justification"
    ws = wb[target["_sheet"]]
    row = int(target["_row"])
    assert ws.cell(row, 21).value == "delivery"
    assert ws.cell(row, 22).value == "procurement"
    assert ws.cell(row, 23).value == "Customs hold"
    wb.close()

    database.init_db()
    excel_service.invalidate()
    client = TestClient(app)
    token = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"}).json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}
    got = client.get(f"/api/work-orders/{rid}", headers=headers)
    assert got.status_code == 200
    item = got.json()["item"]
    assert item["delay_kind"] == "delivery"
    assert item["delay_justification"] == "Customs hold"
    saved = client.put(
        f"/api/work-orders/{rid}",
        headers=headers,
        json={
            "changes": {
                "delay_kind": "placement",
                "delay_source": "site",
                "delay_justification": "API dual-write",
            },
            "force": True,
        },
    )
    assert saved.status_code == 200, saved.text
    assert saved.json()["item"]["delay_kind"] == "placement"
    excel_service.invalidate()
    reloaded = excel_service.get_by_id(rid)
    assert reloaded["delay_kind"] == "placement"
    assert reloaded["delay_justification"] == "API dual-write"
    extra = database.get_record_extra(rid)
    assert extra["delay_kind"] == "placement"
    assert extra["delay_justification"] == "API dual-write"
