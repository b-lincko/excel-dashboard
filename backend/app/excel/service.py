from __future__ import annotations

import hashlib
import os
import re
import shutil
import tempfile
import threading
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Optional

from filelock import FileLock, Timeout
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.formula import ArrayFormula

from .. import database
from ..config import AppConfig, load_config, norm_header
from ..dates import format_date, parse_date

DUE_OFFSETS = {
    "direct cash": 3,
    "local po": 5,
    "international": 10,
    "service": 10,
    "consumable": 2,
    "emergency": 0,
    "alternative": 10,
    "under warranty": 10,
}

DELAY_FIELDS = ("delay_kind", "delay_source", "delay_justification")


def _temp_xlsx(directory: Path) -> Path:
    fd, name = tempfile.mkstemp(suffix=".xlsx", dir=directory)
    os.close(fd)
    return Path(name)


def _is_formula(value: Any) -> bool:
    if isinstance(value, ArrayFormula):
        return True
    return isinstance(value, str) and value.startswith("=")


def _cell_plain(value: Any) -> Any:
    if _is_formula(value):
        return None
    return value


class ExcelUnavailable(Exception):
    pass


class ExcelLocked(Exception):
    pass


class SyncConflict(Exception):
    def __init__(self, message: str, current: Optional[dict] = None):
        super().__init__(message)
        self.current = current


class ExcelService:
    def __init__(self) -> None:
        self._cache: Optional[list[dict[str, Any]]] = None
        self._headers: list[str] = []
        self._mtime: Optional[float] = None
        self._fingerprint: str = ""
        self._stale: bool = False
        self._last_error: Optional[str] = None
        self._delay_columns_ready: bool = False
        self._lock = threading.RLock()

    def cfg(self) -> AppConfig:
        return load_config()

    def excel_path(self) -> Path:
        return Path(self.cfg().excel_path)

    def lock_path(self) -> Path:
        p = self.excel_path()
        return p.with_suffix(p.suffix + ".lock")

    def backup_dir(self) -> Path:
        d = Path(self.cfg().backup_dir)
        d.mkdir(parents=True, exist_ok=True)
        return d

    def available(self) -> bool:
        p = self.excel_path()
        return p.exists() and p.is_file()

    def mtime(self) -> Optional[float]:
        p = self.excel_path()
        if not p.exists():
            return None
        return p.stat().st_mtime

    def mtime_iso(self) -> Optional[str]:
        mt = self.mtime()
        if mt is None:
            return None
        return datetime.fromtimestamp(mt).strftime("%Y-%m-%d %H:%M:%S")

    def fingerprint(self) -> str:
        p = self.excel_path()
        if not p.exists():
            return ""
        h = hashlib.sha256()
        h.update(str(p.stat().st_mtime_ns).encode())
        h.update(str(p.stat().st_size).encode())
        return h.hexdigest()[:16]

    def sync_token(self) -> str:
        return self.fingerprint()

    def data_sheets(self, wb: Workbook) -> list[str]:
        cfg = self.cfg()
        names = [n for n in (cfg.worksheets or []) if n in wb.sheetnames]
        if names:
            return names
        if cfg.worksheet_name in wb.sheetnames:
            return [cfg.worksheet_name]
        return [wb.sheetnames[0]] if wb.sheetnames else []

    def site_label(self, sheet_name: str) -> str:
        return self.cfg().worksheet_labels.get(sheet_name) or sheet_name

    def record_id(self, sheet_name: str, row_number: int) -> str:
        return f"{self.site_label(sheet_name)}:{row_number}"

    def map_row(
        self,
        raw: dict[str, Any],
        row_number: int,
        sheet_name: str,
        cfg: Optional[AppConfig] = None,
        mapping: Optional[dict[str, str]] = None,
        fields: Optional[list[str]] = None,
        site: Optional[str] = None,
    ) -> dict[str, Any]:
        cfg = cfg or self.cfg()
        mapping = mapping or cfg.mapping.excel_to_internal()
        fields = fields or list(cfg.mapping.model_dump().keys())
        site = site or self.site_label(sheet_name)
        internal: dict[str, Any] = {}
        for excel_col, value in raw.items():
            field = mapping.get(norm_header(excel_col))
            if field:
                internal[field] = self._normalize_value(field, value)
        internal["_row"] = row_number
        internal["_sheet"] = sheet_name
        internal["_site"] = site
        internal["record_id"] = f"{site}:{row_number}"
        if not internal.get("department"):
            internal["department"] = site
        for field in fields:
            internal.setdefault(field, "")
        self._apply_due_date(internal)
        return internal

    def _due_offsets(self) -> dict[str, int]:
        cfg = self.cfg()
        offsets = dict(DUE_OFFSETS)
        extra = getattr(cfg, "due_offsets", None) or {}
        for key, days in extra.items():
            try:
                offsets[str(key).strip().lower()] = int(days)
            except (TypeError, ValueError):
                continue
        return offsets

    def _apply_due_date(self, rec: dict[str, Any]) -> None:
        created = parse_date(rec.get("created_date"))
        ptype = str(rec.get("work_type") or "").strip().lower()
        offsets = self._due_offsets()
        default_days = int(getattr(self.cfg(), "due_offset_default_days", 14) or 14)
        if ptype in offsets:
            days = offsets[ptype]
        else:
            days = default_days
        rec["_due_offset_days"] = days
        rec["_due_purchase_type"] = ptype
        if not created:
            stored = parse_date(rec.get("due_date"))
            rec["due_date"] = format_date(stored, with_time=False) if stored else ""
            return
        rec["due_date"] = (created + timedelta(days=days)).strftime("%Y-%m-%d")
        rec["_due_computed"] = True

    def _normalize_value(self, field: str, value: Any) -> Any:
        value = _cell_plain(value)
        if value is None:
            return ""
        if field.endswith("_date") or field in {
            "created_date",
            "scheduled_date",
            "due_date",
            "completion_date",
            "closed_date",
            "last_updated",
        }:
            dt = parse_date(value)
            return format_date(dt) if dt else ""
        if isinstance(value, float) and value == int(value):
            return str(int(value))
        if isinstance(value, int) and field == "work_order_id":
            return str(value)
        return str(value).strip() if isinstance(value, str) else value

    def _load_workbook(self, data_only: bool = False, read_only: bool = False):
        path = self.excel_path()
        if not path.exists():
            raise ExcelUnavailable("Excel file is currently unavailable.")
        try:
            return load_workbook(
                path,
                data_only=data_only,
                read_only=read_only,
                keep_vba=(path.suffix == ".xlsm" and not read_only),
            )
        except PermissionError as exc:
            raise ExcelLocked(
                "Excel file is currently being used by another process. Changes cannot be saved until the file becomes available."
            ) from exc
        except Exception as exc:
            raise ExcelUnavailable(f"Excel file could not be opened: {exc}") from exc

    def _sheet(self, wb: Workbook, name: Optional[str] = None):
        name = name or self.cfg().worksheet_name
        if name in wb.sheetnames:
            return wb[name]
        sheets = self.data_sheets(wb)
        return wb[sheets[0]]

    def _read_headers(self, ws) -> list[str]:
        header_row = self.cfg().header_row
        headers: list[str] = []
        for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row)):
            if cell.value is None:
                headers.append(f"Column{cell.column}")
            else:
                headers.append(str(cell.value))
        while headers and headers[-1].startswith("Column"):
            headers.pop()
        return headers

    def _read_sheet_records(self, ws, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
        cfg = self.cfg()
        header_row = cfg.header_row
        start = cfg.data_start_row
        mapping = cfg.mapping.excel_to_internal()
        fields = list(cfg.mapping.model_dump().keys())
        site = self.site_label(sheet_name)
        headers: list[str] = []
        id_field_header = None
        records: list[dict[str, Any]] = []
        empty_streak = 0
        max_col = 40
        for idx, row in enumerate(ws.iter_rows(min_row=header_row, max_col=max_col), start=header_row):
            if idx == header_row:
                headers = []
                for col_i, cell in enumerate(row, start=1):
                    if cell.value is None:
                        headers.append(f"Column{col_i}")
                    else:
                        headers.append(str(cell.value))
                while headers and headers[-1].startswith("Column"):
                    headers.pop()
                max_col = max(len(headers), 1)
                for h in headers:
                    if mapping.get(norm_header(h)) == "work_order_id":
                        id_field_header = h
                        break
                continue
            if idx < start:
                continue
            raw: dict[str, Any] = {}
            empty = True
            for header, cell in zip(headers, row):
                raw[header] = cell.value
                if _cell_plain(cell.value) not in (None, ""):
                    empty = False
            if empty:
                empty_streak += 1
                if empty_streak > 80:
                    break
                continue
            empty_streak = 0
            wo_val = raw.get(id_field_header) if id_field_header else None
            if wo_val in (None, ""):
                continue
            records.append(
                self.map_row(raw, idx, sheet_name, cfg=cfg, mapping=mapping, fields=fields, site=site)
            )
        return headers, records

    def invalidate(self) -> None:
        with self._lock:
            self._cache = None
            self._mtime = None
            self._fingerprint = ""
            self._stale = False
            self._last_error = None
            self._delay_columns_ready = False

    def _serve_cache(self, err: Optional[str] = None) -> list[dict[str, Any]]:
        if self._cache is None:
            raise ExcelUnavailable(err or "Excel file is currently unavailable.")
        self._stale = True
        self._last_error = err
        return self._cache

    def load(self, force: bool = False) -> list[dict[str, Any]]:
        with self._lock:
            mt = self.mtime()
            if not force and self._cache is not None and mt == self._mtime:
                self._stale = False
                self._last_error = None
                return self._cache
            if not self.available():
                return self._serve_cache("Excel file is currently unavailable.")
            try:
                wb = self._load_workbook(data_only=False, read_only=True)
            except (ExcelLocked, ExcelUnavailable, PermissionError, OSError) as exc:
                if self._cache is not None:
                    return self._serve_cache(str(exc))
                if isinstance(exc, (ExcelLocked, ExcelUnavailable)):
                    raise
                raise ExcelUnavailable(f"Excel file could not be opened: {exc}") from exc
            try:
                all_records: list[dict[str, Any]] = []
                headers: list[str] = []
                mapping_exc = self.cfg().mapping.internal_to_excel()
                needed = [norm_header(mapping_exc.get(f, "")) for f in DELAY_FIELDS]
                delay_ready_all = True
                for sheet_name in self.data_sheets(wb):
                    ws = wb[sheet_name]
                    hdrs, recs = self._read_sheet_records(ws, sheet_name)
                    if hdrs:
                        headers = hdrs
                    present = {norm_header(h) for h in hdrs}
                    if not all(n in present for n in needed if n):
                        delay_ready_all = False
                    all_records.extend(recs)
            finally:
                wb.close()
            self._headers = [norm_header(h) for h in headers]
            self._delay_columns_ready = delay_ready_all and bool(headers)
            self._cache = all_records
            self._mtime = mt
            self._fingerprint = self.fingerprint()
            self._stale = False
            self._last_error = None
            database.set_sync_meta("last_read", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            database.set_sync_meta("mtime", self.mtime_iso() or "")
            database.set_sync_meta("token", self._fingerprint)
            database.set_sync_meta("count", str(len(all_records)))
            return all_records

    def headers(self) -> list[str]:
        self.load()
        return list(self._headers)

    def delay_columns_ready(self) -> bool:
        if self._cache is None:
            try:
                self.load()
            except (ExcelUnavailable, ExcelLocked):
                return False
        return bool(self._delay_columns_ready)

    def get_all(self, force: bool = False) -> list[dict[str, Any]]:
        return list(self.load(force=force))

    def get_by_id(self, wo_id: str) -> Optional[dict[str, Any]]:
        wo_id = str(wo_id)
        recs = self.load()
        for rec in recs:
            if str(rec.get("record_id")) == wo_id:
                return rec
        for rec in recs:
            if str(rec.get("work_order_id")) == wo_id:
                return rec
        return None

    def unique_values(self, field: str) -> list[str]:
        values = set()
        for rec in self.load():
            v = rec.get(field)
            if v not in (None, ""):
                values.add(str(v))
        return sorted(values, key=lambda s: s.lower())

    def lists(self) -> dict[str, list[str]]:
        return {}

    def create_backup(self, reason: str = "write") -> Optional[Path]:
        src = self.excel_path()
        if not src.exists():
            return None
        day = datetime.now().strftime("%Y-%m-%d")
        ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        dest_dir = self.backup_dir() / day
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest = dest_dir / f"{src.stem}_{ts}_{reason}{src.suffix}"
        shutil.copy2(src, dest)
        database.set_sync_meta("last_backup", str(dest))
        return dest

    def list_backups(self, limit: int = 50) -> list[dict[str, Any]]:
        items = []
        root = self.backup_dir()
        if not root.exists():
            return items
        for p in sorted(root.rglob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True):
            items.append(
                {
                    "path": str(p),
                    "name": p.name,
                    "size": p.stat().st_size,
                    "modified": datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                }
            )
            if len(items) >= limit:
                break
        return items

    def restore_backup(self, backup_path: str) -> None:
        src = Path(backup_path)
        if not src.exists():
            raise FileNotFoundError("Backup file not found")
        self.create_backup(reason="pre_restore")
        with self._file_lock():
            shutil.copy2(src, self.excel_path())
        self.invalidate()

    def _file_lock(self, timeout: float = 90.0):
        return FileLock(str(self.lock_path()), timeout=timeout)

    def _validate_saved(self, path: Path) -> None:
        wb = load_workbook(path, read_only=True, data_only=False)
        try:
            if not wb.sheetnames:
                raise ValueError("Workbook has no worksheets after save")
        finally:
            wb.close()

    def _atomic_replace(self, tmp_path: Path) -> None:
        dest = self.excel_path()
        self._validate_saved(tmp_path)
        os.replace(tmp_path, dest)

    def _formula_header_set(self) -> set[str]:
        return {norm_header(c) for c in self.cfg().formula_columns}

    def _ensure_mapped_headers(self, ws, headers: list[str]) -> list[str]:
        """Append missing mapped columns at the end of the header row. Never insert/shift."""
        mapping = self.cfg().mapping.internal_to_excel()
        existing = {norm_header(h) for h in headers}
        header_row = self.cfg().header_row
        next_col = len(headers) + 1
        new_headers = list(headers)
        for _field, excel_col in mapping.items():
            nh = norm_header(excel_col)
            if not nh or nh in existing:
                continue
            ws.cell(header_row, next_col).value = excel_col
            new_headers.append(excel_col)
            existing.add(nh)
            next_col += 1
        return new_headers

    def _write_record_to_sheet(self, ws, rec: dict[str, Any], headers: list[str], row_number: int) -> None:
        mapping = self.cfg().mapping.internal_to_excel()
        skip = self._formula_header_set()
        skip.add(norm_header(mapping.get("due_date", "")))
        header_index = {norm_header(h): i + 1 for i, h in enumerate(headers)}
        for field, excel_col in mapping.items():
            if not excel_col or field not in rec:
                continue
            nh = norm_header(excel_col)
            if not nh or nh in skip:
                continue
            col_idx = header_index.get(nh)
            if not col_idx:
                continue
            cell = ws.cell(row=row_number, column=col_idx)
            if _is_formula(cell.value):
                continue
            value = rec.get(field)
            if value in ("", None):
                cell.value = None
            else:
                if field.endswith("_date"):
                    dt = parse_date(value)
                    cell.value = dt if dt else value
                else:
                    cell.value = value

    def _copy_row_formulas(self, ws, from_row: int, to_row: int, headers: list[str]) -> None:
        skip = self._formula_header_set()
        for col_idx, header in enumerate(headers, 1):
            if norm_header(header) not in skip:
                continue
            src = ws.cell(from_row, col_idx).value
            dest_coord = f"{get_column_letter(col_idx)}{to_row}"
            if isinstance(src, ArrayFormula):
                text = src.text.replace(str(from_row), str(to_row))
                ws.cell(to_row, col_idx).value = ArrayFormula(ref=dest_coord, text=text)
            elif isinstance(src, str) and src.startswith("="):
                ws.cell(to_row, col_idx).value = src.replace(str(from_row), str(to_row))

    def _next_id(self, records: list[dict[str, Any]], sheet_name: str) -> str:
        label = self.site_label(sheet_name)
        if label == "F5":
            max_n = 0
            for rec in records:
                if rec.get("_sheet") != sheet_name:
                    continue
                wo = str(rec.get("work_order_id") or "")
                m = re.search(r"(\d+)$", wo)
                if m:
                    max_n = max(max_n, int(m.group(1)))
            return f"LKF5-{max_n + 1:04d}"
        year = datetime.now().year
        prefix = f"{self.cfg().id_prefix}-{year}-"
        max_n = 0
        for rec in records:
            wo = str(rec.get("work_order_id") or "")
            if wo.startswith(prefix):
                try:
                    max_n = max(max_n, int(wo.split("-")[-1]))
                except ValueError:
                    continue
        return f"{prefix}{max_n + 1:06d}"

    def _next_row(self, ws, headers: list[str]) -> int:
        mapping = self.cfg().mapping.excel_to_internal()
        id_col = 2
        for i, h in enumerate(headers, 1):
            if mapping.get(norm_header(h)) == "work_order_id":
                id_col = i
                break
        max_row = self.cfg().data_start_row - 1
        start = self.cfg().data_start_row
        empty = 0
        for r in range(start, ws.max_row + 1):
            if ws.cell(r, id_col).value not in (None, ""):
                max_row = r
                empty = 0
            else:
                empty += 1
                if empty > 80:
                    break
        return max_row + 1

    def _locate(self, recs: list[dict[str, Any]], wo_id: str) -> dict[str, Any]:
        wo_id = str(wo_id)
        target = next((r for r in recs if str(r.get("record_id")) == wo_id), None)
        if not target:
            target = next((r for r in recs if str(r.get("work_order_id")) == wo_id), None)
        if not target:
            raise KeyError(f"Work order {wo_id} was not found in Excel")
        return target

    def update_record(
        self,
        wo_id: str,
        changes: dict[str, Any],
        username: str,
        sync_token: Optional[str] = None,
        force: bool = False,
    ) -> dict[str, Any]:
        if not self.available():
            raise ExcelUnavailable("Excel file is currently unavailable.")
        try:
            lock = self._file_lock()
            lock.acquire()
        except Timeout as exc:
            raise ExcelLocked(
                "Excel file is currently being used by another process. Changes cannot be saved until the file becomes available."
            ) from exc
        try:
            current_token = self.fingerprint()
            if sync_token and not force and sync_token != current_token:
                current = self.get_by_id(wo_id)
                raise SyncConflict(
                    "The Excel workbook has changed since you last loaded it. Review the latest data before saving.",
                    current=current,
                )
            self.create_backup(reason="update")
            wb = self._load_workbook()
            try:
                all_records: list[dict[str, Any]] = []
                sheet_headers: dict[str, list[str]] = {}
                for sheet_name in self.data_sheets(wb):
                    hdrs, recs = self._read_sheet_records(wb[sheet_name], sheet_name)
                    sheet_headers[sheet_name] = hdrs
                    all_records.extend(recs)
                target = self._locate(all_records, wo_id)
                old = dict(target)
                allowed = set(self.cfg().mapping.model_dump().keys())
                for k, v in changes.items():
                    if k.startswith("_") or k in {"record_id", "department"}:
                        continue
                    if k in allowed:
                        target[k] = v if v is not None else ""
                ws = wb[target["_sheet"]]
                headers = self._ensure_mapped_headers(ws, sheet_headers[target["_sheet"]])
                sheet_headers[target["_sheet"]] = headers
                self._write_record_to_sheet(ws, target, headers, int(target["_row"]))
                tmp = _temp_xlsx(self.excel_path().parent)
                wb.save(tmp)
            finally:
                wb.close()
            try:
                self._atomic_replace(tmp)
            except Exception:
                if tmp.exists():
                    tmp.unlink(missing_ok=True)
                raise
            self.invalidate()
            rid = old.get("record_id")
            updated = self.get_by_id(rid) or self.get_by_id(wo_id)
            assert updated is not None
            self._audit_diff(username, str(updated.get("work_order_id") or wo_id), old, updated)
            database.set_sync_meta("last_write", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            database.set_sync_meta("last_write_user", username)
            return updated
        except PermissionError as exc:
            raise ExcelLocked(
                "Excel file is currently being used by another process. Changes cannot be saved until the file becomes available."
            ) from exc
        finally:
            try:
                lock.release()
            except Exception:
                pass

    def create_record(self, data: dict[str, Any], username: str) -> dict[str, Any]:
        if not self.available():
            raise ExcelUnavailable("Excel file is currently unavailable.")
        try:
            lock = self._file_lock()
            lock.acquire()
        except Timeout as exc:
            raise ExcelLocked(
                "Excel file is currently being used by another process. Changes cannot be saved until the file becomes available."
            ) from exc
        try:
            self.create_backup(reason="create")
            wb = self._load_workbook()
            try:
                site = str(data.get("department") or data.get("_site") or data.get("_sheet") or "")
                sheet_name = None
                labels = self.cfg().worksheet_labels
                for sn, lab in labels.items():
                    if site in {sn, lab}:
                        sheet_name = sn
                        break
                if not sheet_name:
                    sheet_name = self.data_sheets(wb)[0]
                ws = wb[sheet_name]
                headers, sheet_recs = self._read_sheet_records(ws, sheet_name)
                all_records = list(sheet_recs)
                wo_id = str(data.get("work_order_id") or "").strip()
                if not wo_id:
                    wo_id = self._next_id(all_records, sheet_name)
                rec: dict[str, Any] = {k: "" for k in self.cfg().mapping.model_dump().keys()}
                rec.update({k: v for k, v in data.items() if not k.startswith("_")})
                rec["work_order_id"] = wo_id
                rec["created_date"] = rec.get("created_date") or datetime.now().strftime("%Y-%m-%d %H:%M")
                rec["status"] = rec.get("status") or "OPEN"
                rec["_raw"] = {}
                headers = self._ensure_mapped_headers(ws, headers)
                row_number = self._next_row(ws, headers)
                rec["_row"] = row_number
                rec["_sheet"] = sheet_name
                template_row = max(self.cfg().data_start_row, row_number - 1)
                self._copy_row_formulas(ws, template_row, row_number, headers)
                self._write_record_to_sheet(ws, rec, headers, row_number)
                tmp = _temp_xlsx(self.excel_path().parent)
                wb.save(tmp)
            finally:
                wb.close()
            try:
                self._atomic_replace(tmp)
            except Exception:
                if tmp.exists():
                    tmp.unlink(missing_ok=True)
                raise
            self.invalidate()
            created = self.get_by_id(self.record_id(sheet_name, row_number)) or self.get_by_id(wo_id)
            assert created is not None
            database.add_audit(username, "create", work_order_id=wo_id, details="Created material request")
            database.set_sync_meta("last_write", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            return created
        except PermissionError as exc:
            raise ExcelLocked(
                "Excel file is currently being used by another process. Changes cannot be saved until the file becomes available."
            ) from exc
        finally:
            try:
                lock.release()
            except Exception:
                pass

    def delete_record(self, wo_id: str, username: str) -> None:
        if not self.available():
            raise ExcelUnavailable("Excel file is currently unavailable.")
        try:
            lock = self._file_lock()
            lock.acquire()
        except Timeout as exc:
            raise ExcelLocked(
                "Excel file is currently being used by another process. Changes cannot be saved until the file becomes available."
            ) from exc
        try:
            self.create_backup(reason="delete")
            wb = self._load_workbook()
            try:
                all_records = []
                for sheet_name in self.data_sheets(wb):
                    _, recs = self._read_sheet_records(wb[sheet_name], sheet_name)
                    all_records.extend(recs)
                target = self._locate(all_records, wo_id)
                ws = wb[target["_sheet"]]
                ws.delete_rows(int(target["_row"]), 1)
                tmp = _temp_xlsx(self.excel_path().parent)
                wb.save(tmp)
            finally:
                wb.close()
            try:
                self._atomic_replace(tmp)
            except Exception:
                if tmp.exists():
                    tmp.unlink(missing_ok=True)
                raise
            self.invalidate()
            database.add_audit(username, "delete", work_order_id=str(wo_id), details="Deleted material request")
            database.set_sync_meta("last_write", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        except PermissionError as exc:
            raise ExcelLocked(
                "Excel file is currently being used by another process. Changes cannot be saved until the file becomes available."
            ) from exc
        finally:
            try:
                lock.release()
            except Exception:
                pass

    def _audit_diff(self, username: str, wo_id: str, old: dict, new: dict) -> None:
        fields = self.cfg().mapping.model_dump().keys()
        for field in fields:
            if field in {"last_updated", "due_date"}:
                continue
            ov = "" if old.get(field) is None else str(old.get(field))
            nv = "" if new.get(field) is None else str(new.get(field))
            if ov != nv:
                database.add_audit(
                    username,
                    "update",
                    work_order_id=wo_id,
                    field=field,
                    old_value=ov,
                    new_value=nv,
                )

    def replace_from_bytes(self, content: bytes, username: str, filename: str = "upload.xlsx") -> dict[str, Any]:
        if not content or len(content) < 100:
            raise ValueError("The uploaded file is empty or too small to be an Excel workbook.")
        dest = self.excel_path()
        dest.parent.mkdir(parents=True, exist_ok=True)
        tmp = _temp_xlsx(dest.parent)
        tmp.write_bytes(content)
        try:
            self._validate_saved(tmp)
        except Exception as exc:
            tmp.unlink(missing_ok=True)
            raise ValueError(f"That file could not be opened as Excel: {exc}") from exc
        try:
            lock = self._file_lock()
            lock.acquire()
        except Timeout as exc:
            tmp.unlink(missing_ok=True)
            raise ExcelLocked(
                "Excel file is currently being used by another process. Changes cannot be saved until the file becomes available."
            ) from exc
        try:
            if dest.exists():
                self.create_backup(reason="upload")
            os.replace(tmp, dest)
            self.invalidate()
            records = self.load(force=True)
            database.add_audit(username, "upload", details=f"Uploaded workbook {filename} ({len(records)} rows)")
            database.set_sync_meta("last_write", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            database.set_sync_meta("last_write_user", username)
            return self.status()
        except PermissionError as exc:
            raise ExcelLocked(
                "Excel file is currently being used by another process. Changes cannot be saved until the file becomes available."
            ) from exc
        finally:
            try:
                lock.release()
            except Exception:
                pass
            if tmp.exists():
                tmp.unlink(missing_ok=True)

    def ping(self) -> dict[str, Any]:
        """Cheap live check: stat the file and reload only when it actually changed."""
        file_token = self.fingerprint()
        if file_token and file_token != self._fingerprint:
            try:
                self.load()
            except (ExcelLocked, ExcelUnavailable, OSError):
                pass
        cached = self._cache is not None
        stale = bool(file_token and self._fingerprint and file_token != self._fingerprint) or self._stale
        return {
            "available": self.available(),
            "mtime": self.mtime_iso(),
            "sync_token": self._fingerprint or file_token,
            "file_token": file_token,
            "record_count": len(self._cache or []),
            "stale": stale,
            "synchronized": cached,
            "error": None if cached else (self._last_error or "Excel file is currently unavailable."),
            "warning": self._last_error if cached and stale else None,
        }

    def status(self) -> dict[str, Any]:
        available = self.available()
        err = None
        try:
            records = self.load() if available or self._cache is not None else []
        except ExcelUnavailable as e:
            records = list(self._cache or [])
            err = str(e)
        except ExcelLocked as e:
            records = list(self._cache or [])
            err = str(e)
        live = self.ping()
        return {
            "available": available,
            "path": str(self.excel_path()),
            "worksheet": ", ".join(self.cfg().worksheets or [self.cfg().worksheet_name]),
            "mtime": self.mtime_iso(),
            "sync_token": live.get("sync_token") or (self.fingerprint() if available else ""),
            "record_count": len(records),
            "last_read": database.get_sync_meta("last_read"),
            "last_write": database.get_sync_meta("last_write"),
            "last_write_user": database.get_sync_meta("last_write_user"),
            "last_backup": database.get_sync_meta("last_backup"),
            "synchronized": bool(records),
            "stale": live.get("stale") or bool(err and records),
            "error": None if records else err,
            "warning": err if err and records else live.get("warning"),
            "headers": self._headers if (available or self._headers) else [],
        }

    def reconcile_overlay(self, username: str = "sync") -> dict[str, Any]:
        """Keep Excel and SQLite delay/supplier data in both directions.

        - Append Delay Type / Source / Justification headers if missing (no column shift).
        - Newly added empty columns are filled from SQLite (migration of overlay notes).
        - Once columns exist, Excel is source of truth including blanks.
        - Supplier names from Excel are added to the SQLite catalog.
        """
        if not self.available():
            raise ExcelUnavailable("Excel file is currently unavailable.")
        try:
            lock = self._file_lock()
            lock.acquire()
        except Timeout as exc:
            raise ExcelLocked(
                "Excel file is currently being used by another process. Changes cannot be saved until the file becomes available."
            ) from exc
        wrote = False
        pushed_excel = 0
        pulled_db = 0
        suppliers_added = 0
        new_columns = False
        try:
            extras = database.get_all_record_extras()
            original_delay = {
                rid: {f: str(ex.get(f) or "") for f in DELAY_FIELDS} for rid, ex in extras.items()
            }
            known_suppliers = {s["name"].lower() for s in database.list_suppliers()}
            wb = self._load_workbook()
            try:
                sheet_headers: dict[str, list[str]] = {}
                all_records: list[dict[str, Any]] = []
                for sheet_name in self.data_sheets(wb):
                    ws = wb[sheet_name]
                    hdrs, recs = self._read_sheet_records(ws, sheet_name)
                    new_hdrs = self._ensure_mapped_headers(ws, hdrs)
                    if new_hdrs != hdrs:
                        wrote = True
                        new_columns = True
                    sheet_headers[sheet_name] = new_hdrs
                    all_records.extend(recs)
                dirty_ids: set[str] = set()
                for rec in all_records:
                    rid = str(rec.get("record_id") or "")
                    extra = extras.get(rid) or {}
                    for field in DELAY_FIELDS:
                        excel_val = str(rec.get(field) or "").strip()
                        db_val = str(extra.get(field) or "").strip()
                        if new_columns and not excel_val and db_val:
                            rec[field] = db_val
                            extra[field] = db_val
                            extras[rid] = extra
                            dirty_ids.add(rid)
                            pushed_excel += 1
                        else:
                            rec[field] = excel_val
                            if excel_val != db_val:
                                extra[field] = excel_val
                                extras[rid] = extra
                                pulled_db += 1
                    name = str(rec.get("supplier") or "").strip()
                    if name and name.lower() not in known_suppliers:
                        database.add_supplier(name, created_by=username)
                        known_suppliers.add(name.lower())
                        suppliers_added += 1
                if dirty_ids:
                    wrote = True
                    by_sheet: dict[str, list[dict[str, Any]]] = {}
                    for rec in all_records:
                        if str(rec.get("record_id") or "") in dirty_ids:
                            by_sheet.setdefault(rec["_sheet"], []).append(rec)
                    for sheet_name, recs in by_sheet.items():
                        ws = wb[sheet_name]
                        headers = sheet_headers[sheet_name]
                        for rec in recs:
                            self._write_record_to_sheet(ws, rec, headers, int(rec["_row"]))
                if wrote:
                    self.create_backup(reason="reconcile")
                    tmp = _temp_xlsx(self.excel_path().parent)
                    wb.save(tmp)
                else:
                    tmp = None
            finally:
                wb.close()
            if wrote and tmp is not None:
                try:
                    self._atomic_replace(tmp)
                except Exception:
                    if tmp.exists():
                        tmp.unlink(missing_ok=True)
                    raise
            self.invalidate()
            records = self.load(force=True)
            for rec in records:
                rid = str(rec.get("record_id") or "")
                extra = extras.get(rid) or {}
                payload = {f: str(rec.get(f) or "") for f in DELAY_FIELDS}
                previous = original_delay.get(rid) or {f: "" for f in DELAY_FIELDS}
                wo = str(rec.get("work_order_id") or extra.get("work_order_id") or "")
                if payload != previous:
                    database.upsert_record_extra(
                        rid,
                        username,
                        work_order_id=wo,
                        delay_kind=payload["delay_kind"],
                        delay_source=payload["delay_source"],
                        delay_justification=payload["delay_justification"],
                    )
            database.set_sync_meta("last_reconcile", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            return {
                "wrote_excel": wrote,
                "pushed_to_excel": pushed_excel,
                "pulled_to_db": pulled_db,
                "suppliers_added": suppliers_added,
                "record_count": len(records),
            }
        except PermissionError as exc:
            raise ExcelLocked(
                "Excel file is currently being used by another process. Changes cannot be saved until the file becomes available."
            ) from exc
        finally:
            try:
                lock.release()
            except Exception:
                pass


excel_service = ExcelService()
