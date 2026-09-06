from __future__ import annotations

import csv
import io
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import HRFlowable, SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether

from .domain import annotate, is_closed, is_open, is_overdue, matches_filters
from .config import load_config
from .excel.service import excel_service
from .stats import filtered, group_by, kpis, reasons

COLUMNS = [
    ("work_order_id", "IM Work Order #"),
    ("created_date", "MR Received Date"),
    ("description", "Required Material Details"),
    ("department", "Site"),
    ("location", "Location"),
    ("assigned_to", "Assigned To"),
    ("priority", "Priority"),
    ("status", "Status"),
    ("due_date", "Due Date"),
    ("completion_date", "Completion Date"),
    ("closed_date", "Closing Date"),
    ("delay_reason", "Delay Reason"),
    ("remarks", "Remarks"),
    ("work_type", "Work Type"),
    ("issue", "Issue"),
]


def records_for_report(kind: str, filters: dict[str, Any]) -> list[dict[str, Any]]:
    cfg = load_config()
    recs = filtered(excel_service.get_all(), filters)
    if kind == "open":
        recs = [r for r in recs if is_open(r, cfg)]
    elif kind == "overdue":
        recs = [r for r in recs if is_overdue(r, cfg)]
    elif kind == "closed":
        recs = [r for r in recs if is_closed(r, cfg)]
    elif kind == "delay":
        recs = [r for r in recs if is_open(r, cfg)]
    elif kind in {"daily", "weekly", "monthly", "yearly", "department", "technician"}:
        pass
    recs = [annotate(r, cfg) for r in recs]
    recs.sort(key=lambda r: (r.get("created_date") or ""), reverse=True)
    return recs


def to_csv(records: list[dict[str, Any]]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([label for _, label in COLUMNS])
    for r in records:
        writer.writerow([r.get(k) or "" for k, _ in COLUMNS])
    return buf.getvalue().encode("utf-8-sig")


def to_xlsx(records: list[dict[str, Any]], title: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    header_fill = PatternFill("solid", fgColor="0F3D5E")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="0F3D5E")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws["A2"] = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    for i, (_, label) in enumerate(COLUMNS, 1):
        cell = ws.cell(4, i, label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
    for r_i, rec in enumerate(records, 5):
        for c_i, (key, _) in enumerate(COLUMNS, 1):
            cell = ws.cell(r_i, c_i, rec.get(key) or "")
            cell.border = thin
            cell.font = Font(size=9)
    for i, (key, _) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18 if key != "description" else 40
    k = kpis(records)
    summary = wb.create_sheet("Summary")
    summary["A1"] = "KPI Summary"
    summary["A1"].font = Font(bold=True, size=14)
    row = 3
    for label, val in k.items():
        summary.cell(row, 1, label.replace("_", " ").title())
        summary.cell(row, 2, val)
        row += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf(records: list[dict[str, Any]], title: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("T", parent=styles["Title"], fontSize=16, textColor=colors.HexColor("#0F3D5E"), spaceAfter=6)
    meta = ParagraphStyle("M", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#64748B"))
    cell_style = ParagraphStyle("C", parent=styles["Normal"], fontSize=7, leading=9)
    story = [Paragraph(title, title_style), Paragraph(f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}  ·  {len(records)} records", meta), Spacer(1, 8)]
    k = kpis(records)
    kpi_data = [
        ["Total", "Open", "Closed", "Overdue", "Completion", "Avg Close (days)"],
        [
            str(k["total"]),
            str(k["open"]),
            str(k["closed"]),
            str(k["overdue"]),
            f"{k['completion_rate']}%",
            str(k["average_closing_days"] if k["average_closing_days"] is not None else "—"),
        ],
    ]
    kpi_table = Table(kpi_data, colWidths=[40 * mm] * 6)
    kpi_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F3D5E")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#F1F5F9")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(kpi_table)
    story.append(Spacer(1, 10))

    cols = [
        ("work_order_id", "WO #"),
        ("created_date", "Opened"),
        ("department", "Dept"),
        ("assigned_to", "Assigned"),
        ("priority", "Priority"),
        ("status", "Status"),
        ("due_date", "Due"),
        ("delay_reason", "Reason"),
        ("description", "Description"),
    ]
    header = [label for _, label in cols]
    data = [header]
    for rec in records[:400]:
        row = []
        for key, _ in cols:
            val = str(rec.get(key) or "")
            if key in {"created_date", "due_date"}:
                val = val[:10]
            if key == "description":
                val = val[:80]
            row.append(Paragraph(val.replace("&", "&amp;"), cell_style))
        data.append(row)
    widths = [28 * mm, 22 * mm, 28 * mm, 32 * mm, 20 * mm, 24 * mm, 22 * mm, 36 * mm, 60 * mm]
    table = Table(data, colWidths=widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F3D5E")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(table)
    if len(records) > 400:
        story.append(Spacer(1, 6))
        story.append(Paragraph(f"Showing first 400 of {len(records)} records. Export Excel/CSV for the full set.", meta))
    doc.build(story)
    return buf.getvalue()


def _esc(value: Any) -> str:
    return str(value or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def wo_sheet_pdf(rec: dict[str, Any], attachments: Optional[list[dict[str, Any]]] = None) -> bytes:
    """One-page A4 sheet for site: material, supplier, PO, due, attachments."""
    attachments = attachments or []
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"IM WO {rec.get('work_order_id') or rec.get('record_id')}",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "WOTitle", parent=styles["Title"], fontSize=16, textColor=colors.HexColor("#0F3D5E"), spaceAfter=4, alignment=0
    )
    meta = ParagraphStyle("WOMeta", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#64748B"), spaceAfter=2)
    label = ParagraphStyle("WOLabel", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#64748B"), leading=11)
    value = ParagraphStyle("WOValue", parent=styles["Normal"], fontSize=10, leading=13, textColor=colors.HexColor("#0F172A"))
    wo = _esc(rec.get("work_order_id") or rec.get("record_id"))
    story: list[Any] = [
        Paragraph(f"IM Work Order {wo}", title_style),
        Paragraph(
            f"Printed {datetime.now().strftime('%Y-%m-%d %H:%M')} · Linkco MR · Excel remains the source of truth",
            meta,
        ),
        Spacer(1, 4),
        HRFlowable(width="100%", thickness=1, color=colors.HexColor("#0F3D5E"), spaceAfter=8),
    ]
    rows = [
        ["Site", rec.get("department") or "—", "Status", rec.get("status") or "—"],
        ["Assigned to", rec.get("assigned_to") or "—", "Priority", rec.get("priority") or "—"],
        ["Purchase type", rec.get("work_type") or "—", "Due date", str(rec.get("due_date") or "—")[:16]],
        ["Supplier", rec.get("supplier") or "—", "PO No", rec.get("po_number") or "—"],
        ["Asset", rec.get("location") or "—", "Delivery", rec.get("issue") or rec.get("delay_reason") or "—"],
        ["MR received", str(rec.get("created_date") or "—")[:16], "ETA", str(rec.get("closed_date") or "—")[:16]],
    ]
    table_data = []
    for a_l, a_v, b_l, b_v in rows:
        table_data.append(
            [
                Paragraph(_esc(a_l), label),
                Paragraph(_esc(a_v), value),
                Paragraph(_esc(b_l), label),
                Paragraph(_esc(b_v), value),
            ]
        )
    grid = Table(table_data, colWidths=[28 * mm, 63 * mm, 28 * mm, 63 * mm])
    grid.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(grid)
    story.append(Spacer(1, 10))
    story.append(Paragraph("Required material", label))
    story.append(Paragraph(_esc(rec.get("description") or "—")[:1200], value))
    story.append(Spacer(1, 8))
    story.append(Paragraph("Remarks / notes", label))
    story.append(Paragraph(_esc(rec.get("remarks") or "—")[:1200].replace("\n", "<br/>"), value))
    story.append(Spacer(1, 8))
    story.append(Paragraph("Attachments", label))
    if attachments:
        for item in attachments[:20]:
            note = f" — {_esc(item.get('note'))}" if item.get("note") else ""
            story.append(
                Paragraph(
                    f"• {_esc(item.get('filename'))} ({_esc(item.get('kind') or 'file')}, {_esc(item.get('created_by'))} {_esc(str(item.get('created_at') or '')[:16])}){note}",
                    value,
                )
            )
    else:
        story.append(Paragraph("No files attached in the app.", value))
    story.append(Spacer(1, 10))
    story.append(
        Paragraph(
            "Formulas, SN and due-date columns in Excel are not printed. Take this sheet to site; update the workbook in Linkco MR after the visit.",
            meta,
        )
    )
    doc.build([KeepTogether(story)])
    return buf.getvalue()


def render(kind: str, fmt: str, filters: dict[str, Any]) -> tuple[bytes, str, str]:
    records = records_for_report(kind, filters)
    title_map = {
        "daily": "Daily Work Order Report",
        "weekly": "Weekly Work Order Report",
        "monthly": "Monthly Work Order Report",
        "yearly": "Yearly Work Order Report",
        "open": "Open Work Order Report",
        "overdue": "Overdue Work Order Report",
        "closed": "Closed Work Order Report",
        "delay": "Delay / Issue Report",
        "department": "Department Report",
        "technician": "Technician Report",
    }
    title = title_map.get(kind, "Work Order Report")
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    if fmt == "csv":
        return to_csv(records), f"{kind}_report_{stamp}.csv", "text/csv"
    if fmt == "xlsx":
        return to_xlsx(records, title), f"{kind}_report_{stamp}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if fmt == "pdf":
        return to_pdf(records, title), f"{kind}_report_{stamp}.pdf", "application/pdf"
    raise ValueError(f"Unsupported format {fmt}")
