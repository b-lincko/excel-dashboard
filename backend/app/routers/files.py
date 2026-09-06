from __future__ import annotations

import csv
import io
import re
import uuid
from pathlib import Path
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import load_workbook

from .. import database
from ..config import ATTACHMENTS_DIR, load_config, norm_header
from ..excel.service import ExcelLocked, ExcelUnavailable, excel_service
from ..security import require_permission
from ..stats import invalidate_dash_cache

router = APIRouter(tags=["files"])

ALLOWED_EXT = {".pdf", ".png", ".jpg", ".jpeg", ".webp", ".gif", ".csv", ".xlsx", ".xlsm"}
IMAGE_EXT = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
MAX_BYTES = 15 * 1024 * 1024


def _safe_name(name: str) -> str:
    base = Path(name or "file").name
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", base).strip("._") or "file"
    return cleaned[:180]


def _kind_for(name: str, mime: str) -> str:
    ext = Path(name).suffix.lower()
    if ext == ".pdf" or "pdf" in (mime or ""):
        return "pdf"
    if ext in IMAGE_EXT or (mime or "").startswith("image/"):
        return "screenshot"
    return "file"


def _raise_excel(exc: Exception):
    if isinstance(exc, ExcelUnavailable):
        raise HTTPException(status_code=503, detail=str(exc))
    if isinstance(exc, ExcelLocked):
        raise HTTPException(status_code=423, detail=str(exc))
    raise exc


def _map_row(raw: dict[str, Any]) -> dict[str, Any]:
    cfg = load_config()
    mapping = cfg.mapping.excel_to_internal()
    internal = set(cfg.mapping.model_dump().keys()) | {"record_id"}
    out: dict[str, Any] = {}
    for key, value in raw.items():
        if value in (None, ""):
            continue
        header = norm_header(key)
        slug = header.lower().replace(" ", "_").replace("#", "").replace("/", "_").strip("_")
        field = mapping.get(header)
        if not field and slug in internal:
            field = slug
        if not field and header.lower() in internal:
            field = header.lower()
        if field:
            out[field] = value if not isinstance(value, str) else value.strip()
    return out


def _rows_from_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [_map_row(row) for row in reader]


def _rows_from_xlsx(content: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return []
        headers = [norm_header(h) if h is not None else f"Column{i}" for i, h in enumerate(header_row, 1)]
        out = []
        for row in rows_iter:
            raw = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            mapped = _map_row(raw)
            if mapped:
                out.append(mapped)
        return out
    finally:
        wb.close()


def _text_from_pdf(content: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="PDF support is not installed (pypdf).") from exc
    reader = PdfReader(io.BytesIO(content))
    parts = []
    for page in reader.pages[:40]:
        try:
            parts.append(page.extract_text() or "")
        except Exception:
            continue
    return "\n".join(parts).strip()


def _store_file(record_id: str, filename: str, content: bytes, mime: str, username: str, note: str = "") -> dict[str, Any]:
    if len(content) > MAX_BYTES:
        raise HTTPException(status_code=400, detail="File is larger than 15 MB.")
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXT:
        raise HTTPException(status_code=400, detail="Allowed files: PDF, PNG, JPG, WEBP, GIF, CSV, Excel.")
    rec = excel_service.get_by_id(record_id)
    if not rec:
        raise HTTPException(status_code=404, detail="Work order not found")
    folder = ATTACHMENTS_DIR / re.sub(r"[^A-Za-z0-9._-]+", "_", record_id)
    folder.mkdir(parents=True, exist_ok=True)
    stored = f"{uuid.uuid4().hex}_{_safe_name(filename)}"
    dest = folder / stored
    dest.write_bytes(content)
    item = database.add_attachment(
        record_id=str(rec.get("record_id")),
        filename=_safe_name(filename),
        stored_name=str(dest.relative_to(ATTACHMENTS_DIR)),
        mime=mime or "application/octet-stream",
        size=len(content),
        created_by=username,
        work_order_id=str(rec.get("work_order_id") or ""),
        kind=_kind_for(filename, mime),
        note=note,
    )
    return item


@router.get("/api/work-orders/{wo_id}/files")
def list_files(wo_id: str, user=Depends(require_permission("view"))):
    rec = excel_service.get_by_id(wo_id)
    if not rec:
        raise HTTPException(status_code=404, detail="Work order not found")
    return {"items": database.list_attachments(str(rec.get("record_id")))}


@router.post("/api/work-orders/{wo_id}/files")
async def upload_file(
    wo_id: str,
    file: UploadFile = File(...),
    note: str = Form(""),
    user=Depends(require_permission("edit")),
):
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="The uploaded file is empty.")
    item = _store_file(wo_id, file.filename or "upload.bin", content, file.content_type or "", user["username"], note)
    return {"item": item}


@router.get("/api/files/{attachment_id}")
def download_file(attachment_id: int, user=Depends(require_permission("view"))):
    item = database.get_attachment(attachment_id)
    if not item:
        raise HTTPException(status_code=404, detail="Attachment not found")
    path = ATTACHMENTS_DIR / item["stored_name"]
    if not path.is_file():
        raise HTTPException(status_code=404, detail="File is missing on disk")
    return FileResponse(path, filename=item["filename"], media_type=item.get("mime") or "application/octet-stream")


@router.delete("/api/files/{attachment_id}")
def remove_file(attachment_id: int, user=Depends(require_permission("edit"))):
    item = database.get_attachment(attachment_id)
    if not item:
        raise HTTPException(status_code=404, detail="Attachment not found")
    path = ATTACHMENTS_DIR / item["stored_name"]
    database.delete_attachment(attachment_id)
    try:
        if path.is_file():
            path.unlink()
    except OSError:
        pass
    return {"deleted": True, "id": attachment_id}


@router.get("/api/transfer/export.csv")
def export_csv(user=Depends(require_permission("view"))):
    try:
        records = excel_service.get_all()
    except (ExcelUnavailable, ExcelLocked) as exc:
        _raise_excel(exc)
    cfg = load_config()
    fields = list(cfg.mapping.model_dump().keys())
    headers = ["record_id", *fields]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for rec in records:
        row = {k: rec.get(k, "") for k in headers}
        writer.writerow(row)
    data = buf.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="linkco-mr-export.csv"'},
    )


@router.post("/api/transfer/import")
async def import_file(
    file: UploadFile = File(...),
    user=Depends(require_permission("edit")),
):
    name = (file.filename or "upload").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="The uploaded file is empty.")
    rows: list[dict[str, Any]] = []
    pdf_text = ""
    try:
        if name.endswith(".csv"):
            rows = _rows_from_csv(content)
        elif name.endswith((".xlsx", ".xlsm")):
            rows = _rows_from_xlsx(content)
        elif name.endswith(".pdf"):
            pdf_text = _text_from_pdf(content)
            wo_match = re.search(r"(MR-\d{4}-\d+|LKF5-\d+)", pdf_text, re.I)
            if wo_match:
                rec = excel_service.get_by_id(wo_match.group(1))
                if rec:
                    note = (pdf_text[:500] + "…") if len(pdf_text) > 500 else pdf_text
                    item = _store_file(
                        str(rec["record_id"]),
                        file.filename or "import.pdf",
                        content,
                        "application/pdf",
                        user["username"],
                        note=note,
                    )
                    return {
                        "created": 0,
                        "updated": 0,
                        "skipped": 0,
                        "attached": 1,
                        "attachment": item,
                        "errors": [],
                        "message": f"PDF attached to {rec.get('work_order_id')}.",
                    }
            rows = [
                {
                    "description": pdf_text[:4000] or (file.filename or "Imported PDF"),
                    "remarks": f"Imported from PDF: {file.filename}",
                    "status": "OPEN",
                    "department": "SH5-SH1",
                }
            ]
        else:
            raise HTTPException(status_code=400, detail="Use Excel (.xlsx), CSV or PDF.")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read that file: {exc}") from exc
    try:
        result = excel_service.import_rows(rows, username=user["username"])
    except (ExcelUnavailable, ExcelLocked) as exc:
        _raise_excel(exc)
    invalidate_dash_cache()
    if name.endswith(".pdf") and result.get("created"):
        recs = excel_service.get_all()
        created = recs[-1] if recs else None
        if created:
            try:
                _store_file(
                    str(created["record_id"]),
                    file.filename or "import.pdf",
                    content,
                    "application/pdf",
                    user["username"],
                    note="Source PDF for imported material request",
                )
                result["attached"] = 1
            except HTTPException:
                result["attached"] = 0
    result["sync"] = excel_service.ping()
    return result
