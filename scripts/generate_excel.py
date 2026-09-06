#!/usr/bin/env python3
"""Generate a realistic Work Orders Excel workbook used as the source of truth."""
from __future__ import annotations

import random
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import PieChart, Reference
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.marker import DataPoint as DP

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "work_orders.xlsx"

random.seed(20260905)

HEADERS = [
    "WO No",
    "Date Opened",
    "Scheduled Date",
    "Due Date",
    "Completion Date",
    "Closing Date",
    "Work Type",
    "Description",
    "Issue",
    "Department",
    "Location",
    "Assigned To",
    "Priority",
    "Status",
    "Delay Reason",
    "Remarks",
    "Created By",
    "Last Updated",
]

DEPARTMENTS = [
    "Electrical",
    "Mechanical",
    "HVAC",
    "Plumbing",
    "Instrumentation",
    "Facilities",
    "IT/Networks",
    "Safety",
    "Production",
    "Quality",
]

LOCATIONS = [
    "Plant 1 - Line A",
    "Plant 1 - Line B",
    "Plant 2 - Assembly",
    "Warehouse North",
    "Warehouse South",
    "Building A Floor 1",
    "Building A Floor 2",
    "Building B",
    "Utility Room",
    "Rooftop",
    "Parking Area",
    "Office Block",
    "Loading Dock",
    "Boiler House",
    "Substation",
]

TECHNICIANS = [
    "James Wilson",
    "Maria Garcia",
    "Ahmed Hassan",
    "Sarah Chen",
    "David Okonkwo",
    "Priya Patel",
    "Tomasz Kowalski",
    "Emily Johnson",
    "Carlos Rivera",
    "Liu Wei",
    "Anna Schmidt",
    "Michael Brown",
    "Fatima Al-Sayed",
    "Noah Williams",
]

WORK_TYPES = [
    "Corrective",
    "Preventive",
    "Inspection",
    "Emergency",
    "Installation",
    "Calibration",
    "Safety",
    "Improvement",
]

PRIORITIES = ["Critical", "High", "Medium", "Low"]
STATUSES = [
    "New",
    "Open",
    "Assigned",
    "In Progress",
    "Pending",
    "On Hold",
    "Completed",
    "Closed",
    "Cancelled",
]
DELAY_REASONS = [
    "Waiting for spare parts",
    "Waiting for approval",
    "Waiting for technician",
    "Waiting for customer",
    "Material unavailable",
    "Vendor delay",
    "Technical problem",
    "Access unavailable",
    "Equipment unavailable",
    "Safety issue",
    "Insufficient information",
    "Scheduling issue",
    "Weather delay",
    "Other",
]
CREATED_BY = ["Admin", "Maria Garcia", "James Wilson", "Sarah Chen", "Control Room"]

ASSETS = [
    "Main compressor #2",
    "Conveyor belt C-14",
    "Chiller unit CH-03",
    "Transformer T-1",
    "Boiler feed pump",
    "CNC mill #7",
    "Packaging robot R2",
    "Air handling unit AHU-4",
    "Fire pump FP-1",
    "Cooling tower CT-2",
    "Hydraulic press HP-9",
    "Forklift FL-12",
    "PLC cabinet Line A",
    "Emergency generator G-1",
    "Water treatment skid",
    "Silo rotary valve",
    "Overhead crane CR-3",
    "HVAC rooftop RTU-2",
    "Network switch SW-Core",
    "Lighting panel LP-B2",
    "Steam trap ST-18",
    "Exhaust fan EF-6",
    "Dock leveler DL-4",
    "Access control panel",
    "Pressure vessel PV-5",
]

ISSUES = [
    "Abnormal vibration detected during operation",
    "Overheating above operating limit",
    "Intermittent trip / unexpected shutdown",
    "Leak observed at flange / seal",
    "Noise level exceeded threshold",
    "Failed preventive inspection",
    "Sensor reading out of range",
    "Electrical fault / breaker trip",
    "Corrosion and surface damage",
    "Software / PLC communication loss",
    "Blocked drain / reduced flow",
    "Safety interlock not engaging",
    "Calibration overdue",
    "Physical damage after impact",
    "Performance degradation vs baseline",
]

DESCRIPTIONS = [
    "Investigate and restore {asset} to normal operating condition",
    "Perform scheduled maintenance on {asset}",
    "Replace worn components on {asset}",
    "Emergency response for {asset} failure",
    "Inspect {asset} and report findings",
    "Calibrate instruments associated with {asset}",
    "Install upgrade kit on {asset}",
    "Safety check and lockout verification for {asset}",
    "Corrective repair following production stoppage — {asset}",
    "Root-cause analysis and repair of {asset}",
]

REMARKS_OPEN = [
    "Parts ordered, ETA next week",
    "Waiting on vendor quotation",
    "Access only possible during weekend shutdown",
    "Technician assigned, tools staged",
    "Need engineering review before proceeding",
    "Temporary workaround in place",
    "Requires specialized contractor",
    "Production will not release equipment until Friday",
    "",
    "Follow-up inspection scheduled",
]
REMARKS_CLOSED = [
    "Repaired and tested. Returned to service.",
    "Replaced bearing assembly. Vibration within spec.",
    "Leak repaired, pressure test passed.",
    "Calibration completed, certificates filed.",
    "No fault found after inspection.",
    "Preventive tasks completed as planned.",
    "Cancelled — duplicate of another work order.",
    "Completed during planned outage.",
]


def daterange_random(start: datetime, end: datetime) -> datetime:
    delta = (end - start).days
    return start + timedelta(days=random.randint(0, max(delta, 0)), hours=random.randint(6, 18), minutes=random.randint(0, 59))


def wo_number(year: int, seq: int) -> str:
    return f"WO-{year}-{seq:06d}"


def pick_status(opened: datetime, due: datetime, today: datetime) -> str:
    age = (today - opened).days
    overdue = due < today
    r = random.random()
    if age > 40 and r < 0.72:
        return random.choices(["Closed", "Completed", "Cancelled"], weights=[80, 12, 8])[0]
    if age > 20 and r < 0.55:
        return random.choices(["Closed", "Completed", "In Progress", "Pending"], weights=[60, 10, 18, 12])[0]
    if overdue and r < 0.35:
        return random.choices(["In Progress", "Pending", "On Hold", "Open", "Assigned"], weights=[30, 25, 15, 20, 10])[0]
    return random.choices(
        STATUSES,
        weights=[8, 10, 8, 16, 10, 6, 8, 30, 4],
    )[0]


def build_rows(today: datetime) -> list[list]:
    rows = []
    seq = {2024: 1, 2025: 1, 2026: 1}
    counts = {2024: 92, 2025: 110, 2026: 88}
    year_ranges = {
        2024: (datetime(2024, 1, 8, 8, 0), datetime(2024, 12, 20, 16, 0)),
        2025: (datetime(2025, 1, 6, 8, 0), datetime(2025, 12, 19, 16, 0)),
        2026: (datetime(2026, 1, 5, 8, 0), datetime(2026, 9, 5, 12, 0)),
    }

    for year, n in counts.items():
        start, end = year_ranges[year]
        for _ in range(n):
            opened = daterange_random(start, end)
            if opened > today:
                opened = today - timedelta(days=random.randint(0, 10))
            work_type = random.choice(WORK_TYPES)
            if work_type == "Emergency":
                priority = random.choices(PRIORITIES, weights=[55, 35, 8, 2])[0]
                lead = random.randint(1, 5)
            elif work_type == "Preventive":
                priority = random.choices(PRIORITIES, weights=[2, 15, 53, 30])[0]
                lead = random.randint(7, 30)
            else:
                priority = random.choices(PRIORITIES, weights=[8, 28, 44, 20])[0]
                lead = random.randint(3, 18)

            scheduled = opened + timedelta(days=random.randint(0, 4), hours=random.randint(0, 8))
            due = opened + timedelta(days=lead, hours=random.randint(2, 10))
            status = pick_status(opened, due, today)

            closed_statuses = {"Closed", "Completed", "Cancelled"}
            completion = None
            closing = None
            delay = ""
            if status in closed_statuses:
                duration = random.randint(1, max(lead + random.randint(-2, 12), 2))
                completion = opened + timedelta(days=duration, hours=random.randint(1, 7))
                if completion > today:
                    completion = today - timedelta(hours=random.randint(1, 20))
                if status == "Cancelled":
                    closing = completion
                    delay = random.choice(["Other", "Insufficient information", "Scheduling issue", ""])
                else:
                    closing = completion + timedelta(hours=random.randint(0, 30))
                    if closing > today:
                        closing = today
                    if closing.date() > due.date():
                        delay = random.choice(DELAY_REASONS)
                    else:
                        delay = random.choice(["", "", "", "Scheduling issue"])
                remarks = random.choice(REMARKS_CLOSED)
            else:
                if due < today or status in {"Pending", "On Hold"}:
                    delay = random.choices(DELAY_REASONS, weights=[18, 12, 10, 6, 8, 10, 9, 6, 6, 5, 4, 4, 1, 1])[0]
                else:
                    delay = random.choice(["", "", random.choice(DELAY_REASONS)])
                remarks = random.choice(REMARKS_OPEN)
                if status in {"Completed"}:
                    completion = opened + timedelta(days=random.randint(1, 10))
                    if completion > today:
                        completion = today - timedelta(hours=4)

            asset = random.choice(ASSETS)
            desc = random.choice(DESCRIPTIONS).format(asset=asset)
            issue = random.choice(ISSUES)
            dept = random.choice(DEPARTMENTS)
            loc = random.choice(LOCATIONS)
            tech = random.choice(TECHNICIANS)
            created_by = random.choice(CREATED_BY)
            last_updated = closing or completion or (opened + timedelta(days=random.randint(0, max((today - opened).days, 0))))
            if last_updated > today:
                last_updated = today

            wo = wo_number(year, seq[year])
            seq[year] += 1

            def fmt(dt):
                return dt.strftime("%Y-%m-%d %H:%M") if dt else ""

            rows.append(
                [
                    wo,
                    opened.strftime("%Y-%m-%d %H:%M"),
                    scheduled.strftime("%Y-%m-%d %H:%M"),
                    due.strftime("%Y-%m-%d %H:%M"),
                    fmt(completion) if completion else "",
                    fmt(closing) if closing else "",
                    work_type,
                    desc,
                    issue,
                    dept,
                    loc,
                    tech,
                    priority,
                    status,
                    delay,
                    remarks,
                    created_by,
                    last_updated.strftime("%Y-%m-%d %H:%M"),
                ]
            )

    rows.sort(key=lambda r: r[1])
    # Re-number in chronological order within year to keep IDs stable-looking
    counters = {2024: 1, 2025: 1, 2026: 1}
    for r in rows:
        y = int(r[0].split("-")[1])
        r[0] = wo_number(y, counters[y])
        counters[y] += 1
    return rows


def style_header(ws, cols):
    fill = PatternFill("solid", fgColor="0F3D5E")
    font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )
    ws.row_dimensions[1].height = 28
    for col in range(1, cols + 1):
        cell = ws.cell(1, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = thin


def apply_row_style(ws, row_idx, n_cols, status):
    thin = Border(
        left=Side(style="thin", color="E6EAF0"),
        right=Side(style="thin", color="E6EAF0"),
        top=Side(style="thin", color="E6EAF0"),
        bottom=Side(style="thin", color="E6EAF0"),
    )
    fills = {
        "Closed": PatternFill("solid", fgColor="E8F6EE"),
        "Completed": PatternFill("solid", fgColor="E7F1FB"),
        "Cancelled": PatternFill("solid", fgColor="F1F5F9"),
        "Critical_open": PatternFill("solid", fgColor="FDECEC"),
        "On Hold": PatternFill("solid", fgColor="FFF6E5"),
        "Pending": PatternFill("solid", fgColor="FFF6E5"),
    }
    fill = None
    if status in fills:
        fill = fills[status]
    for c in range(1, n_cols + 1):
        cell = ws.cell(row_idx, c)
        cell.border = thin
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        if fill:
            cell.fill = fill


def autosize(ws, max_width=42):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col[:80]:
            val = "" if cell.value is None else str(cell.value)
            length = max(length, min(len(val), max_width))
        ws.column_dimensions[letter].width = min(max(length + 3, 12), max_width)


def main():
    today = datetime(2026, 9, 5, 12, 0)
    rows = build_rows(today)
    OUT.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ---- WorkOrders ----
    ws = wb.active
    ws.title = "WorkOrders"
    for i, h in enumerate(HEADERS, 1):
        ws.cell(1, i, h)
    for r_i, row in enumerate(rows, 2):
        for c_i, val in enumerate(row, 1):
            ws.cell(r_i, c_i, val)
        apply_row_style(ws, r_i, len(HEADERS), row[13])
    style_header(ws, len(HEADERS))
    last_row = 1 + len(rows)
    last_col = get_column_letter(len(HEADERS))
    table = Table(displayName="WorkOrdersTable", ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Data validations
    status_dv = DataValidation(type="list", formula1="=Lists!$A$2:$A$10", allow_blank=True)
    status_dv.error = "Select a valid status"
    status_dv.errorTitle = "Invalid Status"
    prio_dv = DataValidation(type="list", formula1="=Lists!$B$2:$B$5", allow_blank=True)
    type_dv = DataValidation(type="list", formula1="=Lists!$C$2:$C$9", allow_blank=True)
    delay_dv = DataValidation(type="list", formula1="=Lists!$D$2:$D$15", allow_blank=True)
    for dv, col in ((status_dv, "N"), (prio_dv, "M"), (type_dv, "G"), (delay_dv, "O")):
        dv.add(f"{col}2:{col}{last_row + 200}")
        ws.add_data_validation(dv)

    # Conditional formatting for Critical priority
    ws.conditional_formatting.add(
        f"M2:M{last_row}",
        FormulaRule(formula=['$M2="Critical"'], fill=PatternFill("solid", fgColor="FECACA"), font=Font(color="7F1D1D", bold=True)),
    )
    autosize(ws)
    ws.column_dimensions["H"].width = 48
    ws.column_dimensions["I"].width = 40
    ws.column_dimensions["P"].width = 40
    ws.sheet_properties.tabColor = "0F3D5E"

    # ---- Lists ----
    lists = wb.create_sheet("Lists")
    list_headers = ["Status", "Priority", "Work Type", "Delay Reason", "Department", "Location", "Assigned To"]
    list_data = [STATUSES, PRIORITIES, WORK_TYPES, DELAY_REASONS, DEPARTMENTS, LOCATIONS, TECHNICIANS]
    for i, h in enumerate(list_headers, 1):
        lists.cell(1, i, h)
    style_header(lists, len(list_headers))
    for c, values in enumerate(list_data, 1):
        for r, v in enumerate(values, 2):
            lists.cell(r, c, v)
            lists.cell(r, c).font = Font(name="Calibri", size=10)
    autosize(lists)
    lists.sheet_properties.tabColor = "1D6A96"
    lists.freeze_panes = "A2"

    # ---- Summary (formulas — must be preserved) ----
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Work Order Summary (Excel formulas — do not overwrite)"
    summary["A1"].font = Font(name="Calibri", bold=True, size=14, color="0F3D5E")
    summary.merge_cells("A1:C1")
    summary["A3"] = "Metric"
    summary["B3"] = "Value"
    style_header(summary, 2)
    metrics = [
        ("Total Work Orders", f"=COUNTA(WorkOrders[WO No])"),
        ("Closed", f'=COUNTIF(WorkOrders[Status],"Closed")'),
        ("Completed", f'=COUNTIF(WorkOrders[Status],"Completed")'),
        ("Cancelled", f'=COUNTIF(WorkOrders[Status],"Cancelled")'),
        ("In Progress", f'=COUNTIF(WorkOrders[Status],"In Progress")'),
        ("Pending", f'=COUNTIF(WorkOrders[Status],"Pending")'),
        ("On Hold", f'=COUNTIF(WorkOrders[Status],"On Hold")'),
        ("Open", f'=COUNTIF(WorkOrders[Status],"Open")'),
        ("New", f'=COUNTIF(WorkOrders[Status],"New")'),
        ("Assigned", f'=COUNTIF(WorkOrders[Status],"Assigned")'),
        ("Critical Priority", f'=COUNTIF(WorkOrders[Priority],"Critical")'),
    ]
    for i, (name, formula) in enumerate(metrics, 4):
        summary.cell(i, 1, name).font = Font(name="Calibri", size=11)
        summary.cell(i, 2, formula).font = Font(name="Calibri", size=11, bold=True)
        summary.cell(i, 2).alignment = Alignment(horizontal="center")
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 18
    summary["A16"] = "Note: These formulas are calculated by Excel. The dashboard computes equivalent statistics independently from the WorkOrders table."
    summary["A16"].alignment = Alignment(wrap_text=True)
    summary.merge_cells("A16:F16")
    summary.sheet_properties.tabColor = "C9A227"

    # ---- Instructions ----
    info = wb.create_sheet("Instructions")
    info["A1"] = "Work Order Workbook — Source of Truth"
    info["A1"].font = Font(name="Calibri", bold=True, size=16, color="0F3D5E")
    notes = [
        "This workbook is the authoritative dataset for the Work Order Management Dashboard.",
        "Do not rename the WorkOrders sheet or the WO No column — they are used as the unique key.",
        "The dashboard reads and writes the WorkOrders table. Summary formulas are preserved.",
        "Statuses, priorities, work types and delay reasons are defined on the Lists sheet.",
        "Date fields use YYYY-MM-DD HH:MM. Closing Date should be filled when Status is Closed.",
        "Last Updated is maintained by the dashboard when records are edited.",
        "A backup is created automatically before every write from the application.",
    ]
    info.merge_cells("A1:G1")
    for i, n in enumerate(notes, 3):
        info.cell(i, 1, n)
        info.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
        info.cell(i, 1).font = Font(name="Calibri", size=11)
    info.column_dimensions["A"].width = 120
    info.sheet_properties.tabColor = "64748B"

    wb.properties.title = "Work Order Management Source of Truth"
    wb.properties.creator = "WOMS"
    wb.properties.description = "Authoritative work-order dataset for the Work Order Management Dashboard"

    wb.save(OUT)
    print(f"Wrote {len(rows)} work orders to {OUT}")
    from collections import Counter
    print("Status:", Counter(r[13] for r in rows))
    print("Year:", Counter(r[0].split("-")[1] for r in rows))


if __name__ == "__main__":
    main()
